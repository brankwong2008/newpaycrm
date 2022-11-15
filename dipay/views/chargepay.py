
import os
from django.utils.safestring import mark_safe
from stark.service.starksite import StarkHandler,Option
from django.shortcuts import render, HttpResponse
from stark.utils.display import PermissionHanlder, get_date_display,get_choice_text
from dipay.utils.displays import ttcopy_display
from django.http import JsonResponse
from dipay.models import PayToCharge,Charge
from dipay.forms.forms import ChargePayModelForm
from django.conf import settings
from django.conf.urls import url
from datetime import datetime

class ChargePayHandler(StarkHandler):

    show_detail_template = "dipay/show_chargepay_detail.html"

    search_list = ['id__icontains', "create_date"]

    search_placeholder = '搜索 付费单号 日期'

    def amount_display(self, obj=None, is_header=None,*args,**kwargs):
        """  美元金额合计显示  """
        if is_header:
           return "金额"
        else:
            return "%s %s" % (obj.currency.icon , obj.amount)

    def related_charges_display(self,obj=None, is_header=None,*args,**kwargs):
        if is_header:
            return "关联账单"
        else:
            queryset = PayToCharge.objects.filter(chargepay_id=obj.pk)
            download_url = self.reverse_url("download",pk=obj.pk)
            bills_download_btn = f"<a href='{download_url}'>下载明细</a>"
            return mark_safe( "   ".join([ str(item.charge.followorder)
                                 +"("+item.currency.icon+str(item.amount or '.') +")  "
                                 for item in queryset]) + bills_download_btn)

    def charge_id_display(self,obj=None, is_header=None,*args,**kwargs):
        if is_header:
            return "付费单号"
        else:

            return  "F%s" % str(obj.pk).zfill(5)

    fields_display = [
        charge_id_display,
        get_date_display("create_date"),
        'bank',
        'forwarder',
        amount_display,
        get_choice_text("status"),
        ttcopy_display,
        related_charges_display,
        "remark",

    ]

    def get_extra_urls(self):
        return [
            url("^download/(?P<pk>\d+)/$", self.wrapper(self.download), name=self.get_url_name('download')),
        ]

    def save_form(self, form, request, is_update=False, *args, **kwargs):
        # 上传水单说明实际付款了，可以更新相关表的状态
        if form.instance.ttcopy:
            # 更新付款表时，如果上传水单，则把付款状态改为已出账
            form.instance.status = 1

            # 更新支付日期为传水单的当日, 这个地方不应该自动，有可能第二天或者第三天传水单，还是应允许用户修改
            # form.instance.create_date = datetime.now()

            # 且同时要把相关联的付费单的状态改变美元已付，人民币已付，或者结清
            currency_code = 1 if form.instance.currency.title == '美元' else 2
            for item in Charge.objects.filter(chargepay=form.instance):
                # 如果currency与费用表中的状态值相等，说明已经更新过了
                if item.status < 3 and item.status != currency_code:
                    item.status += currency_code
                    item.save()
        form.save()


    def get_model_form(self,type=None):
        return ChargePayModelForm

    # 显示一条记录详情
    def show_detail(self, request, pk, *args, **kwargs):
        # print('pk',pk)
        obj = self.model_class.objects.filter(pk=pk).first()
        data_list = []

        edit_detail_url = self.reverse_edit_url(pk=pk)

        return render(request, self.show_detail_template, locals())

    # 下载这笔付款所覆盖的费用单的明细，财务需要打出来做账
    def download(self, request, pk, *args, **kwargs):
        chargepay_obj = self.model_class.objects.filter(pk=pk).first()
        if not chargepay_obj:
            return HttpResponse("付费单号不存在")

        sample_file = os.path.join(settings.MEDIA_ROOT, "charge_list_sample.xlsx")

        # 读入sample_file
        from openpyxl import load_workbook
        wb = load_workbook(sample_file,data_only=True)
        ws = wb.active
        ws["B1"].value = chargepay_obj.forwarder.title
        ws["E1"].value = "F" + str(pk).zfill(5)

        USD_total, CNY_total = 0,0
        for item in PayToCharge.objects.filter(chargepay_id=pk):
            USD_amount,CNY_amount  = 0,0
            if item.currency ==1:
                USD_amount = item.amount
                USD_total += USD_amount
            else:
                CNY_amount = item.amount
                CNY_total += CNY_amount
            BL_date = item.charge.BL_date.strftime("%Y-%m-%d")
            remark = item.charge.remark
            order_number = str(item.charge.followorder)
            row = [BL_date,remark,order_number,USD_amount or "-", CNY_amount or "-"]
            ws.append(row)
        ws.append(["合计","","", USD_total or "", CNY_total] or "")

        file_name = "F%s.xlsx" % (str(pk).zfill(5))
        file_path = os.path.join(settings.MEDIA_ROOT, file_name)
        wb.save(file_path)

        with open(file_path, 'rb') as f:
            try:
                response = HttpResponse(f)
                response['Content-Type'] = 'application/octet-stream'
                response['Content-Disposition'] = 'attachment;filename="%s"' % (file_name)

                return response
            except Exception as e:
                print(e)
                return HttpResponse("下载失败")