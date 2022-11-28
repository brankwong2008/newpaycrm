import os
from stark.service.starksite import StarkHandler
from stark.utils.display import PermissionHanlder
from django.conf.urls import url
from django.shortcuts import render,HttpResponse, redirect
from openpyxl import load_workbook


class CustomerHandler(PermissionHanlder,StarkHandler):

    verify_similarity_list = ['title','shortname']

    page_title = "客户"

    fields_display = ['id','title','owner']

    def get_per_page(self):
        return 10

    search_list = ['title__icontains','owner__username__icontains']
    search_placeholder = '搜索 客户名'

    def get_queryset_data(self,request,*args,**kwargs):
        if request.user.username == 'brank':
            return self.model_class.objects.all()
        if request.user.department == 8:
            return self.model_class.objects.all()
        elif request.user.department == 1:
            return self.model_class.objects.filter(owner = request.user)
    def get_extra_urls(self):
        return [
            url("^upload/$", self.wrapper(self.upload_customer), name=self.get_url_name('upload_customer')),
        ]

        # 删除一条记录
    def upload_customer(self,request, *args, **kwargs):
        print(request.POST, request.FILES)
        if request.POST:
            customer_file = request.FILES.get('customer_file')
            print('yes upload',customer_file)
            # 存储文件
            file_path = os.path.join("media/", customer_file.name)

            # 存入media文件夹
            with open(file_path, "wb") as f:
                for line in customer_file:
                    f.write(line)

            # 读取excel文件
            excel_file = load_workbook(file_path)
            ws = excel_file.active
            count = 0

            # 遍历每行，获取每一行信息
            customer_list = []
            for row in ws.iter_rows(2):
                title = row[0].value
                owner = row[2].value
                print(title,owner)
                customer_obj = self.model_class(title=title,owner_id=owner, shortname=title)
                customer_list.append(customer_obj)
            self.model_class.objects.bulk_create(customer_list)
            return HttpResponse('upload successfully...')

        return render(request,'dipay/upload_customer.html',locals())
