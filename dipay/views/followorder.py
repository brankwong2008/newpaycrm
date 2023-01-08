import os
from django.shortcuts import HttpResponse, redirect, render, reverse
from django.db.models.functions import TruncMonth
from datetime import datetime
from django.conf import settings
from django.http import JsonResponse
from django.utils.safestring import mark_safe
from stark.service.starksite import StarkHandler, Option
from stark.utils.display import get_date_display, get_choice_text, PermissionHanlder, checkbox_display,checkbox_display_func
from dipay.utils.displays import status_display, info_display, save_display, \
    follow_date_display, order_number_display, sales_display, port_display, goods_display, customer_display, \
    term_display, amount_display, confirm_date_display, rcvd_amount_blance_display,basic_info_display,\
    customer_goods_port_display,amount_rvcd_collect_display,book_info_display, more_tag_display

from django.db.models import ForeignKey

from dipay.forms.forms import AddApplyOrderModelForm, EditFollowOrderModelForm
from django.db import models
from django.conf.urls import url
from dipay.models import ApplyOrder, Pay2Orders, ApplyRelease, UserInfo, FollowOrder
from decimal import Decimal
from dipay.utils.order_updates import order_payment_update
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment


class FollowOrderHandler(PermissionHanlder, StarkHandler):
    # 添加按钮
    has_add_btn = False
    detail_fields_display = ['order', 'load_port', 'discharge_port', 'ETD', 'ETA', 'book_info', 'load_info',
                             'produce_info', 'sales_remark', ]
    page_title = "跟单"
    show_list_template = 'dipay/show_follow_order_list.html'
    order_by_list = ['-order__confirm_date', '-order__sequence', ]

    search_list = ['order__order_number__icontains', 'order__goods__icontains', 'order__customer__shortname__icontains',
                   'order__customer__title__icontains','book_info__icontains','load_info__icontains']
    search_placeholder = '搜 订单号/客户/货物/装箱/订舱'


    # 加入一个组合筛选框, default是默认筛选的值，必须是字符串
    option_group = [
        Option(field='status', is_multi=False, default='1'),
        Option(field='salesman', filter_param={'roles__title': '外销员'}, verbose_name='业务'),
        # Option(field='depart'),
    ]

    # 按月筛选的字段列表
    def get_time_search(self,time_query):
        year_month = ""
        if time_query:
            year_month = time_query.split("__")[1]
        month_list = FollowOrder.objects.exclude(ETD__isnull=True). \
            annotate(month=TruncMonth("ETD")).values("month").distinct().order_by("-month")
        options = []
        is_chozen = ""
        for each in month_list:
            time_item = each["month"].strftime("%Y-%m")
            if time_item == year_month:
                options.append({"val":time_item, "selected":"selected"})
                is_chozen = "time-search-chozen"
            else:
                options.append({"val": time_item, "selected": ""})


        time_search = [{
            "title": "ETD",
            "field": "ETD",
            "options": options,
            "active": is_chozen,
        }, ]

        return time_search

    # 批量排产
    def batch_to_produce(self, request, *args, **kwargs):
        pk_list = request.POST.getlist('pk')
        for pk in pk_list:
            followorder_obj = self.model_class.objects.filter(pk=pk).first()
            followorder_obj.status = 1
            followorder_obj.save()
        return redirect(self.reverse_list_url())

    batch_to_produce.text = '批量排产'

    # 检查发票金额
    def check_amount(self, request, *args, **kwargs):
        print("enter check amont, ", request.POST.getlist('pk'))
        pk_list = request.POST.getlist('pk')
        for pk in pk_list:
            followorder_obj = self.model_class.objects.filter(pk=pk).first()
            print(followorder_obj.order.amount_check, not followorder_obj.order.amount_check )
            followorder_obj.order.amount_check = not followorder_obj.order.amount_check
            followorder_obj.order.save()
        return JsonResponse({"status":True, "msg":"检查发票金额成功"})

    check_amount.text = "金额核查"

    # 批量拆分订单的 选项卡
    def batch_split_order(self, request, *args, **kwargs):
        pk_list = request.POST.getlist('pk')
        if len(pk_list) > 1:
            msg = '只能选择一个订单'
            return render(request, 'dipay/msg_after_submit.html', locals())
        data_list = []
        for pk in pk_list:
            follow_order_obj = self.model_class.objects.filter(pk=pk).first()
            order_pk = follow_order_obj.order.pk
            # 更新订单应收已收记录
            order_payment_update(order_obj=follow_order_obj.order)

            # 订单号只取前五位，如J3879-1, order_number: J3879
            order_number = follow_order_obj.order.order_number[:5]
            customer = follow_order_obj.order.customer.shortname
            sub_sequence_num = follow_order_obj.order.sub_sequence
            if sub_sequence_num == 0:
                sub_sequence_num = 1
            sub_sequence = mark_safe(
                "<input name='sub_sequence' type='text' value='%s' readonly>" % (sub_sequence_num))
            goods = mark_safe(
                '<input name="goods" type="text" value="%s"> ' % (follow_order_obj.order.goods))
            amount = mark_safe(
                '<input class="split-amount" name="amount"  type="text" value="%s" onblur="rectAmount(this)" total_amount="%s" > ' % (
                    follow_order_obj.order.amount,follow_order_obj.order.amount))

            rcvd_amount = mark_safe(
                '<input class="split-rcvd-amount" name="rcvd_amount"  type="text" value="%s" onblur="rectRcvdAmount(this)" total_rcvd_amount="%s"> ' % (
                    follow_order_obj.order.rcvd_amount, follow_order_obj.order.rcvd_amount))

            data_list.append([order_number, customer, sub_sequence, goods, amount, rcvd_amount])
            amount = mark_safe(
                '<input class="split-amount" name="amount"  type="text" value="0" onblur="rectAmount(this)" total_amount="%s" >'% follow_order_obj.order.amount )
            rcvd_amount = mark_safe(
                '<input class="split-rcvd-amount" name="rcvd_amount"  type="text" value="0"  onblur="rectRcvdAmount(this)" total_rcvd_amount="%s"> '% follow_order_obj.order.rcvd_amount )

            sub_sequence = mark_safe(
                "<input name='sub_sequence' type='text' value='%s' readonly>" % (sub_sequence_num + 1))

            data_list.append([order_number, customer, sub_sequence, goods, amount, rcvd_amount])

            return render(request, 'dipay/split_order.html', locals())

    # 拆分订单的view
    def split_record(self, request, pk, *args, **kwargs):

        fields_list = ['sub_sequence', 'goods', 'amount']
        length = len(request.POST.getlist(fields_list[0]))
        # 初始化数据列表，含n个字典
        data_list = [{} for i in range(length)]
        # 整理成列表套字典的格式，方便后续添加到ORM
        for field in fields_list:
            for n, item in enumerate(request.POST.getlist(field)):
                data_list[n][field] = item

        order_obj = ApplyOrder.objects.filter(pk=pk).first()
        pay2order_queryset = Pay2Orders.objects.filter(order=order_obj).order_by('amount')
        followorder_obj = order_obj.followorder
        count = 0
        splited_order_list = []
        order_list = []
        for i in range(length):
            # 更新已有订单
            for key, val in data_list[i].items():
                setattr(order_obj, key, val)
            order_obj.order_number = "%s%s-%s" % (
                order_obj.get_order_type_display(), order_obj.sequence, order_obj.sub_sequence)
            if i == 0:
                try:
                    order_obj.save()
                except Exception as e:
                    msg = '订单号可能重复，请检查。 错误内容：%s' % e
                    return render(request, 'dipay/msg_after_submit.html', locals())
            else:
                try:
                    # 获取拆分定金金额
                    dist_amount = Decimal(request.POST.getlist('rcvd_amount')[i])
                    # 拆分订单并分配金额
                    self.dist_deposit_order(request, pay2order_queryset, order_obj,followorder_obj, dist_amount)

                except Exception as e:
                    msg = '订单号可能重复，请检查。 错误内容：%s' % e
                    return render(request, 'dipay/msg_after_submit.html', locals())

            # 拆分订单列表，为后面更新应收已收做准备
            splited_order_list.append([order_obj.pk, order_obj.order_number])
            count += 1
            list_url = self.reverse_list_url()
            order_list = [f"<a href='{list_url}?q={item[1][:5]}'> {item[1]} </a>" for item in
                          splited_order_list]

        # 在最后更新订单应收已收
        for each in splited_order_list:
            order_payment_update(order_obj, order_id=each[0])

        msg = mark_safe('成功拆分%s个订单: %s ' % (count, ' '.join(order_list)))
        return render(request, 'dipay/msg_after_submit.html', locals())

    batch_split_order.text = '拆分订单'

    def dist_deposit_order(self,request, pay2order_queryset, order_obj,followorder_obj, dist_amount):
        # 新增拆分的订单, 把id置为None即可
        order_obj.pk = None
        order_obj.save()
        # 创建新跟单记录, 清空ETD  ETA, Status =0
        temp_data = [("pk", None), ("ETD", None), ("ETA", None), ("status", 0),
                     ("order_id", order_obj.pk), ("book_info", "订舱"),
                     ("salesman", order_obj.salesperson)]
        [setattr(followorder_obj, x[0], x[1]) for x in temp_data]
        followorder_obj.save()

        # 处理款项重新分配
        for pay2order_obj in pay2order_queryset:
            # pay2order_queryset 已经按金额由小到大排序
            if pay2order_obj.amount <= dist_amount:
                # 如果该款项小于目标分配，直接分配到-2订单
                pay2order_obj.order_id = order_obj.pk
                dist_amount -= pay2order_obj.amount
                pay2order_obj.save()
            else:
                pay2order_obj.amount -= dist_amount
                pay2order_obj.save()
                # 创建新的关联记录, 给拆分的新订单建立新的关联记录
                new_pay2order = Pay2Orders(order_id=order_obj.pk, amount=dist_amount,
                                           rate=pay2order_obj.rate, payment=pay2order_obj.payment)
                new_pay2order.save()
                # 因为是从小金额开始捋，所以遇到可分配款项大于待分配金额时，就可以跳出循环
                break

    # 申请放单
    def apply_release(self, request, *args, **kwargs):
        pk_list = request.POST.getlist('pk')
        if len(pk_list) != 1:
            return render(request, 'dipay/msg_after_submit.html', {'msg': '请选择一单仅且一单'})

        pk = pk_list[0]
        followorder_obj = self.model_class.objects.filter(pk=pk).first()
        order_obj = followorder_obj.order
        if ApplyRelease.objects.filter(order=order_obj).exists():
            return render(request, 'dipay/msg_after_submit.html', {'msg': '已经申请过，不用重复提交'})

        user = request.user
        verifier = UserInfo.objects.filter(roles__title__contains='财务').first()
        applyrelease_obj = ApplyRelease(apply_date=datetime.now(),
                                        applier=user,
                                        order=order_obj,
                                        verifier=verifier,
                                        )
        applyrelease_obj.save()
        apply_release_url = reverse('stark:dipay_applyrelease_list')
        return redirect(apply_release_url)

    apply_release.text = '申请放单'



    # 批量处理列表
    batch_process_list = [batch_to_produce, batch_split_order, apply_release,check_amount]

    def edit_display(self, obj=None, is_header=False, *args, **kwargs):
        """
        在列表页显示编辑按钮
        :param obj:
        :param is_header:
        :return:
        """
        if is_header:
            return "操作"
        else:
            edit_url = self.reverse_edit_url(pk=obj.id)
            if obj.status == 0:
                return mark_safe("<i class='fa fa-edit'></i>")
            else:
                return mark_safe("<a href='%s'><i class='fa fa-edit'></i></a>" % edit_url)

    def details_display(self, obj=None, is_header=False, *args, **kwargs):
        """
        在列表页显示编辑按钮
        :param obj:
        :param is_header:
        :return:
        """
        if is_header:
            return mark_safe("<span class='hidden-lg'>详情</span>")
        else:
            detail_url = self.reverse_url('show_detail', pk=obj.pk)
            return mark_safe("<a href='%s' class='hidden-lg'><i class='fa fa-caret-down'></i></a>" % detail_url)

    # 跟单列表显示的字段内容
    fields_display = [checkbox_display_func(hidden_xs='hidden-xs'),
                      basic_info_display,
                      customer_goods_port_display,
                       status_display,
                      follow_date_display('ETD', time_format='%m/%d'),
                      follow_date_display('ETA', time_format='%m/%d'),
                      info_display('load_info',hidden_xs='hidden-xs'),
                      book_info_display('book_info',hidden_xs='hidden-xs'),
                      info_display('produce_info',hidden_xs='hidden-xs'),
                      amount_rvcd_collect_display,
                      more_tag_display,
                      details_display,
                      ]

    # 自定义按钮的权限控制
    def get_extra_fields_display(self, request, *args, **kwargs):
        permission_dict = request.session.get(settings.PERMISSION_KEY)
        save_url_name = '%s:%s' % (self.namespace, self.get_url_name('save'))

        if save_url_name in permission_dict:
            return [save_display, ]
        else:
            return []

    def get_queryset_data(self, request, *args, **kwargs):
        if request.user.username == 'brank':
            return self.model_class.objects.all()

        if request.user.roles.all().filter(title='外销员').exists():
            return self.model_class.objects.filter(salesman = request.user)

        return self.model_class.objects.all()

    def get_per_page(self):
        return 20

    def get_model_form(self, handle_type=None):
        return EditFollowOrderModelForm

    def get_extra_urls(self):
        patterns = [
            url("^save/$", self.wrapper(self.save_record), name=self.get_url_name('save')),
            url("^split/(?P<pk>\d+)/$", self.wrapper(self.split_record), name=self.get_url_name('split')),
            url("^show_pay_details/(?P<order_id>\d+)/$", self.wrapper(self.show_pay_details),
                name=self.get_url_name('show_pay_details')),
            url("^neating/$", self.wrapper(self.neating), name=self.get_url_name('neating')),
            url("^tests/$", self.wrapper(self.tests), name=self.get_url_name('tests')),
            url("^download/$", self.wrapper(self.download), name=self.get_url_name('download')),
        ]

        return patterns

    def show_pay_details(self, request, order_id, *args, **kwarg):
        order_obj = ApplyOrder.objects.filter(pk=order_id).first()
        print(1232434,'order obj', order_obj)
        if not order_obj:
            return HttpResponse('订单号不存在')

        if not hasattr(order_obj, 'followorder'):
            return HttpResponse('跟单记录不存在，请先创建该订单跟单记录')

        payment_list = Pay2Orders.objects.filter(order=order_obj).order_by('-payment__create_date')


        title = '固定定金' if order_obj.order_number.startswith('L') else '收款明细'
        # 催款的邮件链接
        if order_obj.customer:
            modal_title = '%s (%s)' % (title, order_obj.customer.shortname)
            if hasattr(order_obj.customer,'email'):
                mail = {}
                mail['email'] = order_obj.customer.email
                mail['subject'] = order_obj.order_number + ' order shipping status and balance '
                ETD = order_obj.followorder.ETD if order_obj.followorder.ETD else 'to be updated'
                ETA = order_obj.followorder.ETA if order_obj.followorder.ETA else 'to be updated'
                mail['content'] = 'Dear ' + '%0A%0C%0A%0C' + \
                                  f"The order {order_obj.order_number} status:" + '%0A%0C%0A%0C' + \
                                  f"ETD: {ETD} " + '%0A%0C' + \
                                  f"ETA: {ETA}" + '%0A%0C%0A%0C' + \
                                  f"Invoice Value   : {order_obj.currency.icon}{order_obj.amount}" + '%0A%0C' + \
                                  f"Deposit Recieved: {order_obj.currency.icon}{order_obj.rcvd_amount}" + '%0A%0C' + \
                                  f"Balance Amount : {order_obj.currency.icon}{order_obj.collect_amount}" + '%0A%0C%0A%0C'
        else:
            modal_title = '%s (%s)' % (title, '--')



        return render(request, 'dipay/order_payment_record.html', locals())

    # 每行数据保存的方法，使用ajax
    def save_record(self, request, *args, **kwargs):
        # print('request.POST',request.POST)
        if request.is_ajax():
            data_dict = request.POST.dict()
            pk = data_dict.pop('pk')
            data_dict.pop('csrfmiddlewaretoken')

            followorder_obj = self.model_class.objects.filter(pk=pk).first()
            if not followorder_obj:
                res = {'status': False, 'msg': 'obj not found'}
            else:
                for item, val in data_dict.items():
                    # 金额要保存到applyorder表
                    print('follow obj, item val',item, val)
                    if item == 'amount':
                        applyorder_obj = followorder_obj.order
                        try:
                            for each in ['$','￥',',']:
                                val = val.replace(each,'')
                            applyorder_obj.amount = Decimal(val)
                            # 同时更新应收款
                            applyorder_obj.save()
                            order_payment_update(order_obj=applyorder_obj)
                        except Exception as e:
                            data_dict['status'] = False
                    # elif item =='shipline':
                    #     followorder_obj.shipline_id = val
                    #     # print('followorder_obj dir:', followorder_obj, dir(followorder_obj._meta.model))
                    #     temp_model = followorder_obj._meta.model
                    #     field_obj = temp_model._meta.get_field(item)
                    #     print(field_obj,type(field_obj))
                    #     if isinstance(field_obj, ForeignKey):
                    #         print('this is foreign Key')
                    #         item = item + '_id'
                    #
                    elif item == 'ETA':
                        followorder_obj.ETA = val
                        followorder_obj.is_notified = False

                    else:
                        temp_model = followorder_obj._meta.model
                        field_obj = temp_model._meta.get_field(item)
                        print(field_obj, type(field_obj))
                        if isinstance(field_obj, ForeignKey):
                            item = item + '_id'
                        setattr(followorder_obj, item, val)
                        # 如果follow_order完结，更新applyorder的status
                        if followorder_obj.status == '4' or followorder_obj.status == 4:
                            followorder_obj.order.status = 3
                            followorder_obj.order.save()

                followorder_obj.save()
                data_dict['status'] = True
                res = data_dict
            return JsonResponse(res)

    # 预留的接口，用户批量整理数据资料
    def neating(self, request, *args, **kwargs):
        count = 0
        for obj in self.model_class.objects.all():
            obj.salesman = obj.order.salesperson
            obj.save()
            count += 1
        return HttpResponse('整理成功%s条数据' % count)

    # 预留的接口，用户批量整理数据资料
    def tests(self, request, *args, **kwargs):
        count = 0
        return render(request, 'dipay/tests.html')


    # 下载跟单表  （需要用户指定下载的时间段）
    def download(self,request, *args, **kwargs):
        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")
        print(request.GET)
        # 返回提示用户输入起止时间
        if not start_date:
            download_url = self.reverse_url("download")
            return render(request, "dipay/download_followup.html",locals())

        filter_param = {"order__confirm_date__gte":start_date,"order__confirm_date__lte":end_date }
        if request.user.roles.filter(title__in=["外销员"]).exists() and request.user.username !="brank":
            filter_param["salesman"]=request.user


        # 返回下载文件
        queryset = FollowOrder.objects.filter(**filter_param).order_by("order__confirm_date","order__order_number")   # followorder_queryset
        file_name = "follow_up_%s_%s.xlsx" % (request.user.username, datetime.now().strftime("%Y-%m-%d-%H%M%S"))
        file_path = os.path.join("media/followorder", file_name)
        f = self.make_orderdownload_file(queryset, file_path)
        try:
            response = HttpResponse(f)
            response['Content-Type'] = 'application/octet-stream'
            response['Content-Disposition'] = 'attachment;filename="%s"' % (file_name)
            f.close()
            return response
        except Exception as e:
            print(e)
            f.close()
            return HttpResponse("下载失败")


    def make_orderdownload_file(self,queryset,file_path):
        """
        将queryset中数据写入跟单表下载excel文件
        :param queryset: 已经筛选好的数据 followorder_queryset
        :return: file句柄
        """
        print("make_orderdownload_file", file_path, queryset)
        wb = load_workbook("media/followorder/followup_template.xlsx")
        ws = wb.active
        for each in queryset:
            data_dict = {}
            data_dict["A"] = each.order.order_number  # 订单号
            data_dict["B"] = each.order.salesperson.nickname[0] if  each.order.salesperson else "-" #  #B:  业务员
            data_dict["C"] = ""  #   C:  跟单
            data_dict["D"] = each.get_status_display()  #   D:  跟单状态
            data_dict["E"] = each.order.customer.shortname if each.order.customer else "-" #  E： 客户
            data_dict["F"] = ""  #  F： 特殊
            data_dict["G"] = each.order.goods  #  G： Goods
            data_dict["H"] = each.order.get_term_display()  #  H： Term
            data_dict["I"] = each.discharge_port or "-"  #  I:  Ports
            data_dict["J"] = each.order.confirm_date.strftime("%Y-%m-%d") if each.order.confirm_date else "-"  #  J: 下单日
            data_dict["K"] = each.ETD.strftime("%Y-%m-%d") if each.ETD else "-"  #  K： ETD
            data_dict["L"] = each.ETA.strftime("%Y-%m-%d") if each.ETA else "-"   #  L:  ETA
            data_dict["M"] = each.shipline.shortname if each.shipline else "-"  #  M： 船公司
            data_dict["N"] = each.load_info  #  N:  装箱info
            data_dict["O"] = each.book_info  #  O： 订舱info
            data_dict["P"] = each.order.currency.icon if each.order.currency else "-"  #   P： 币种
            data_dict["Q"] = each.order.amount  #  Q： 发票金额
            data_dict["R"] = ""  #  R: 尾款收齐日
            pay2order_queryset = Pay2Orders.objects.filter(order=each.order).order_by("payment__create_date")
            if each.get_status_display()=="完成" and pay2order_queryset.exists() :
                data_dict["R"] = pay2order_queryset.last().payment.create_date.strftime("%Y-%m-%d")  # R: 尾款收齐日

            row = [item for item in data_dict.values()]
            # 收款的处理 S列开始都是收款记录
            for item in pay2order_queryset:
                row.append(item.amount*item.rate)
                payment_detail = "%s %s%s" % (
                    item.payment.create_date.strftime("%Y/%m/%d"),
                    item.payment.currency.icon,
                    item.payment.amount )
                row.append(payment_detail)

            ws.append(row)
        # 设置单元格格式

        font = Font(name='Arial',
                    size=8,
                    color='FF000000',
                    bold=False,
                    italic=False,
                    vertAlign='baseline', )
        font_bold = Font(name='Arial',
                    size=8,
                    color='FF000000',
                    bold=True,
                    italic=False,
                    vertAlign='baseline',
                       )
        align_center = Alignment(horizontal="center",vertical="center")
        align_right = Alignment(horizontal="right",vertical="center")
        align_left = Alignment(horizontal="left",vertical="center")
        for row in ws.iter_rows(2):
            for cell in row:
                cell.font = font
                if cell.column_letter in ["P",]:
                    cell.alignment = align_right
                elif cell.column_letter in ["Q",]:
                    cell.alignment = align_left
                elif cell.column_letter in ["S","U","W"]:
                    cell.font = font_bold
                    cell.alignment  = align_center

        wb.save(file_path)

        f = open(file_path,'rb')
        return f

