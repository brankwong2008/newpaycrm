import os
from datetime import datetime
from django.shortcuts import HttpResponse, redirect, render, reverse
from django.conf import settings
from django.utils.safestring import mark_safe
from stark.service.starksite import StarkHandler, Option
from stark.utils.display import get_date_display, get_choice_text, PermissionHanlder
from dipay.forms.forms import AddApplyOrderModelForm, EditApplyOrderModelForm, ConfirmApplyOrderModelForm, \
    ManualAddApplyOrderModelForm
from dipay.models import CurrentNumber, Customer, FollowOrder, ApplyOrder, Currency, Inwardpay, Bank, Pay2Orders, \
    UserInfo, Payer
from openpyxl import load_workbook
from django.conf.urls import url
from django.db import transaction
import re
from decimal import Decimal


class ApplyOrderHandler(PermissionHanlder, StarkHandler):
    # 每页显示记录数
    def get_per_page(self):
        return 10

    # 自定义列表，外键字段快速添加数据，在前端显示加号
    popup_list = ['customer', ]

    # 加入一个组合筛选框
    option_group = [
        Option(field='status'),
        # Option(field='depart'),
    ]

    # 模糊搜索
    search_list = ['customer__title__icontains', 'goods__icontains', 'order_number__icontains']
    search_placeholder = '搜索 客户/货品/订单号'

    # 添加按钮
    has_add_btn = False

    # 排序字段
    order_by_list = ['-sequence', ]

    def rcvd_amount_display(self, obj=None, is_header=False, *args, **kwargs):
        """
               显示币种和金额
               :param obj:
               :param is_header:
               :return:
               """
        if is_header:
            return '已收款'
        else:
            payments_url_name = "%s:%s" % (self.namespace, 'dipay_pay2orders_list')
            payments_url = reverse(payments_url_name)+'?order_id=%s' % obj.pk
            return mark_safe("<a href=%s target='_blank'> %s </a>" % (payments_url, obj.rcvd_amount))

    def amount_display(self, obj=None, is_header=False, *args, **kwargs):
        """
               显示币种和金额
               :param obj:
               :param is_header:
               :return:
               """
        if is_header:
            return '发票金额'
        else:
            amount_text = "%s %s" % (obj.currency.icon, obj.amount)
            return amount_text

    def order_number_display(self, obj=None, is_header=False, *args, **kwargs):
        """
               显示币种和金额
               :param obj:
               :param is_header:
               :return:
               """
        if is_header:
            return '发票号'
        else:
            if obj.status == 0:
                return obj.order_number[0] + "____"
            else:
                return obj.order_number

    def status_display(self, obj=None, is_header=False, *args, **kwargs):
        """
               显示币种和金额
               :param obj:
               :param is_header:
               :return:
               """
        if is_header:
            return '状态'
        else:
            if obj.status == 0:
                return mark_safe("<span style='color:red'> %s </span>" % obj.get_status_display())
            else:
                return obj.get_status_display()

    def edit_display(self, obj=None, is_header=False, *args, **kwargs):
        """
        在列表页显示编辑按钮
        :param obj:
        :param is_header:
        :return:
        """
        if is_header:
            return "操作"
        else:
            edit_url = self.reverse_edit_url(pk=obj.id)
            if obj.status == 0:
                return mark_safe("<i class='fa fa-edit'></i>")
            else:
                return mark_safe("<a href='%s'><i class='fa fa-edit'></i></a>" % edit_url)

    def to_workshop_display(self, obj=None, is_header=False, *args, **kwargs):
        """
        在列表页显示编辑按钮
        :param obj:
        :param is_header:
        :return:
        """
        if is_header:
            return "下单"
        else:
            if obj.status == 1:
                confirm_url = self.reverse_url('to_workshop', pk=obj.id)
                return mark_safe("<a href='%s'>下单</a>" % confirm_url)
            else:
                return '-'

    fields_display = [get_date_display('create_date'),
                      'salesperson', order_number_display,
                      'customer', 'goods',
                      status_display, to_workshop_display,
                      amount_display, rcvd_amount_display]

    def get_model_form(self, handle_type=None):
        if handle_type == 'add':
            return AddApplyOrderModelForm
        if handle_type == 'edit':
            return EditApplyOrderModelForm
        else:
            return ManualAddApplyOrderModelForm

    # 根据用户筛选数据源
    def get_queryset_data(self, request, *args, **kwargs):
        if request.user.username in ['brank','tony'] :
            return self.model_class.objects.all()
        if request.user:
            return self.model_class.objects.filter(salesperson=request.user, status__lte=3)
        return self.model_class.objects.all()

    # 申请订单号
    def add_list(self, request, *args, **kwargs):
        """ 申请/添加订单号  """
        if request.method == "GET":
            form = self.get_model_form("add")()
            # 限定在客户选项里面，业务员只能看到自己的客户
            choices = [(None, '----------'), ]
            for item in Customer.objects.filter(owner=request.user).values('id', 'title'):
                choices.append((item['id'], item['title']))
            form.fields['customer'].choices = choices

            back_url = self.reverse_list_url()
            return render(request, "dipay/apply_new_order.html", locals())

        if request.method == "POST":
            form = self.get_model_form("add")(data=request.POST)

            if form.is_valid():
                result = self.save_form(form, request, False, *args, **kwargs)

                return result or redirect(self.reverse_list_url(*args, **kwargs))
            else:
                return render(request, self.add_list_template or "stark/apply_new_order.html", locals())

    # 保存表单数据
    def save_form(self, form, request, is_update=False, *args, **kwargs):
        order_type = form.instance.get_order_type_display()
        # order_number 只显示订单类型
        # 编辑情况直接save
        if is_update:
            order_type = form.instance.get_order_type_display()
            sub_sequence = form.instance.sub_sequence
            sequence = str(form.instance.sequence)
            sequence = sequence.zfill(4) if len(sequence) < 4 else sequence
            order_number = "%s%s" % (order_type,sequence)
            if sub_sequence != 0:
                order_number = "%s-%s" % (order_number, sub_sequence)

            form.instance.order_number = order_number
            form.instance.collect_amount = form.instance.amount - form.instance.rcvd_amount
            form.save()
            msg = '订单信息更新成功'
            print('msg:', msg)

            return redirect(self.reverse_list_url())

        # 新增订单的情况下

        # 获取当前最新订单序号，并把新申请订单号置为：最新单号+1
        current_num_obj = CurrentNumber.objects.get(pk=1)
        sequence = current_num_obj.num + 1
        # 检查订单序号是否已经存在，如果存在的话，sequence号顺序后移
        while True:
            order_obj = ApplyOrder.objects.filter(sequence=sequence)
            if order_obj:
                sequence += 1
            else:
                break

        form.instance.sequence = sequence

        form.instance.order_number = "%s%s" % (order_type, form.instance.sequence)
        # 应收初始金额等于订单金额
        form.instance.collect_amount = form.instance.amount

        # 如果是添加新订单，则同步更新最新订单号
        if request.user:
            form.instance.salesperson = request.user
        else:
            return HttpResponse('请先登录')

        form.save()

        # 同步更新最新订单号
        current_num_obj.num = sequence
        current_num_obj.save()

        # 新功能： 自动发申请订单的邮件，节约大家的时间
        send_time = datetime.now().strftime("%Y-%m-%d")
        subject = '申请合同号 %s %s %s' % (form.instance.customer.shortname, form.instance.goods, send_time)
        content = '经理: %0A%0C%0A%0C请审核。' + form.instance.remark
        # 替换特殊符号 & ， 因为&在mailto字符串里面有功能含义
        subject =  subject.replace('&',' and ')
        content =  content.replace('&',' and ')
        mailto = f'<a href="mailto:brank@diligen.cn?subject={subject}&body={content}">点击快速发申请邮件</a>'
        msg = mark_safe('订单号申请提交成功，%s' % mailto)

        return render(request, 'dipay/msg_after_submit.html', {'msg': msg})

    # 自定义路由： 上传，下载，下单，手动新增订单
    def get_extra_urls(self):
        patterns = [
            url("^upload/$", self.wrapper(self.upload), name=self.get_url_name('upload')),
            url("^download/(?P<file_name>.*)/$", self.wrapper(self.download), name=self.get_url_name('download')),
            url("^to_workshop/(?P<pk>\d+)/$", self.wrapper(self.to_workshop), name=self.get_url_name('to_workshop')),
            url("^manual_add/$", self.wrapper(self.manual_add), name=self.get_url_name('manual_add')),
        ]
        return patterns

    #  批量上传订单跟单和收款信息
    def upload(self, request, *args, **kwargs):
        print(request.POST, request.FILES)

        if request.method == "GET":
            download_file_url = self.reverse_url('download', file_name='订单导入模板.xlsx')
            return render(request, 'dipay/upload_orders.html', locals())

        if request.method == "POST":
            # 跟单文件导入
            upload_order_file = request.FILES.get('upload_order')
            # 收款文件导入
            upload_pay_file = request.FILES.get('upload_pay')
            upload_file = upload_pay_file or upload_order_file
            if upload_file is None:
                return HttpResponse('文件不存在')
            print('yes upload', upload_file)
            # 存储文件
            file_path = os.path.join("media/", upload_file.name)

            # 存入media文件夹
            with open(file_path, "wb") as f:
                for line in upload_file:
                    f.write(line)

            # 读取excel文件
            excel_file = load_workbook(file_path, data_only=True)
            ws = excel_file.active
            count = 0

            # 跟单文件的导入
            if upload_order_file:
                count, count1, count2, errors = self.parse_order_file(ws)
                excel_file.close()
                msg = '成功读取 %s 条记录，更新%s条记录，新增%s条记录☺' % (count, count1, count2)
                return render(request, 'dipay/msg_after_submit.html', {"msg": msg, 'errors': errors})

            # 付款文件的导入
            if upload_pay_file:
                count, count1, errors = self.parse_pay_file(ws)
                excel_file.close()

                msg = '成功上传 %s 条分配数据, 收款%s条' % (count, count1)

                return render(request, 'dipay/msg_after_submit.html', {"msg": msg, 'errors': errors})

    # 跟单文件导入的细节处理
    def parse_order_file(self, ws):
        count = 0  # 读入记录数
        count1 = 0  # 更新记录数
        count2 = 0  # 新增记录数
        field_list = [(0, 'order_number'), (1, 'salesperson'), (3, 'status'), (4, 'customer'),
                      (6, 'goods'), (7, 'term'), (8, 'ports'), (9, 'confirm_date'), (10, 'ETD'),
                      (11, 'ETA'), (12, 'load_info'), (13, 'book_info'), (14, 'payterm'), (15, 'amount'),
                      (16, 'deposit'), (17, 'balance_USD'), (18, 'balance_RMB'), (19, 'payment1'), (20, 'payment2')]

        # 准备基础信息源，避免多次撞库
        salespersons = {obj.nickname[0]: obj for obj in UserInfo.objects.all()}
        status_dict = {item[1]: item[0] for item in FollowOrder.follow_choices}
        customers = {obj.title: obj for obj in Customer.objects.all()}
        errors = []

        # 读入并清理数据
        for i in range(4, ws.max_row + 1):
            count += 1
            row = ws[i]
            row_dict = {}
            if not row[0].value:
                continue
            if not re.match(Regex.order_number, row[0].value):
                continue

            pass_flag = False

            # 读取一条数据
            for n, field in field_list:
                row_dict[field] = row[n].value
                if n == 1:
                    row_dict[field] = salespersons.get(row[n].value)
                # 读取状态
                elif n == 3:
                    row_dict[field] = status_dict.get(row[n].value)
                elif n == 4:
                    exists_customer = False
                    if row[n].value:
                        for title, obj in customers.items():
                            if row[n].value[:6] in title:
                                row_dict[field] = obj
                                exists_customer = True
                                break
                    if not exists_customer:
                        row_dict[field] = None
                elif n == 8:
                    if row[n].value:
                        port_list = row[n].value.split('-')
                        if len(port_list) == 1:
                            row_dict['load_port'], row_dict['discharge_port'] = None, row[n].value
                        else:
                            row_dict['load_port'], row_dict['discharge_port'] = port_list[0], port_list[1]
                elif n == 9:
                    # 控制导入时间
                    confirm_date = row[n].value
                    if not confirm_date or confirm_date < datetime(2020, 1, 1):
                        pass_flag = True
                        break
                    else:
                        row_dict[field] = confirm_date
                elif n == 15:
                    amount = row[n].value
                    row_dict[field] = amount
                    if amount and not (isinstance(amount, float) or isinstance(amount, int)):
                        amount_match = re.search(Regex.amount, amount)
                        if amount_match:
                            row_dict[field] = amount_match.group()

            print('行号：%s, content: %s' % (i, row_dict))

            # 创建一条记录
            if not pass_flag:
                d = row_dict
                # 前面已经排除了order_number为空的情况
                order_number = d['order_number'].strip()
                order_obj = ApplyOrder.objects.filter(order_number=order_number).first()

                if order_obj:
                    for field in ['amount', 'confirm_date', 'customer', 'goods', 'salesperson']:
                        if d.get(field):
                            setattr(order_obj, field, d[field])
                        if field == 'customer' and not d.get(field):
                            e = '客户名没有添加到系统'
                            errors.append('错误行号：%s，订单号：%s, 错误原因：%s' % (i, order_number, e))

                    # 如果订单存在，且跟单记录存在，则更新ETD, ETA, load_info, book_info
                    followorder_obj = FollowOrder.objects.filter(order=order_obj).first()
                    if followorder_obj:
                        for field in ['ETA', 'ETD', 'load_info', 'book_info', 'status']:
                            if d.get(field):
                                # 如果表里面的statu落后于CRM中则不更新
                                if field == 'status' and d.get(field) < followorder_obj.status:
                                    continue
                                setattr(followorder_obj, field, d[field])

                    # 如果订单不存在，则创建跟单记录
                    else:
                        followorder_obj = FollowOrder(order=order_obj, salesman=order_obj.salesperson)
                        for field in ['load_port', 'discharge_port', 'ETA', 'ETD', 'load_info', 'book_info', 'status']:
                            if d.get(field):
                                setattr(followorder_obj, field, d[field])
                    try:
                        # 跟单状态4：完成  ， 订单状态3：完结
                        if followorder_obj.status == 4:
                            order_obj.status = 3
                        order_obj.save()
                        followorder_obj.save()
                        count1 += 1
                    except Exception as e:
                        print(e)
                        errors.append('错误行号：%s，订单号：%s, 错误原因：%s' % (i, order_number, e))
                    continue

                # order_obj不存在时，新建
                if not re.match(Regex.order_number, order_number):
                    print(e)
                    errors.append('错误行号：%s，订单号：%s, 错误原因：%s' % (i, order_number, e))
                    continue

                order_type, sequence, sub_sequence = self.parse_order_number(order_number)
                deposit = ' 定金：'
                if d.get('deposit'):
                    deposit = deposit + '%s' % d.get('deposit')
                balance = ' 尾款：'
                for item in ['payment1', 'payment2']:
                    if d.get(item):
                        balance = balance + '%s ' % d.get(item)

                order_obj = ApplyOrder(order_number=order_number, order_type=order_type,
                                       sequence=sequence, sub_sequence=sub_sequence,
                                       status=2, remark=deposit + balance)
                for field in ['confirm_date', 'customer', 'salesperson', 'goods', 'amount']:
                    if d.get(field):
                        setattr(order_obj, field, d[field])
                try:
                    order_obj.save()
                except Exception as e:
                    print(e)
                    errors.append('错误行号：%s，订单号：%s, 错误原因：%s' % (i, order_number, e))
                    continue

                followorder_obj = FollowOrder(order=order_obj)
                followorder_obj.sales_remark = deposit + balance
                for field in ['load_port', 'discharge_port', 'ETA', 'ETD', 'load_info', 'book_info', 'status']:
                    if d.get(field):
                        setattr(followorder_obj, field, d[field])
                try:
                    # 如果已经完结的订单，订单status要更新
                    if followorder_obj.status == 4:
                        order_obj.status = 3
                        order_obj.save()
                    followorder_obj.save()
                    count2 += 1
                except Exception as e:
                    print(e)
                    errors.append('错误行号：%s，订单号：%s, 错误原因：%s' % (i, order_number, e))
                    continue

        return count, count1, count2, errors

    # 收款文件导入的细节处理
    def parse_pay_file(self, ws):
        """
        :param ws: 打开的workbook.activ_sheet
        :return:   处理完的count数，和没有成功导入的行号列表
        pay_list的数据格式
          {  date:2022/5/6  got_amount:23000.00 currency_id 1  related_order: [{'order_number':J3155, amount:$2500]     }
        """
        # 准备基本信息   reference: 收款编号   record_num: 分配记录号
        field_list = [(0, 'create_date'), (2, 'order_number'), (3, 'payer'),
                      (6, 'got_amount'), (7, 'origin_pay'),(9,'dist_amount'), (10,'amount'),(11, 'reference'), (12, 'dist_ref')]
        bank_list = ['稠州', '东亚', '广发', '花旗', '连连']
        banks = {}
        for item in bank_list:
            bank_obj = Bank.objects.filter(title__icontains=item).first()
            if bank_obj:
                banks[item] = bank_obj.pk
        currencies = {}
        for item in ['美元', '加元', '人民币']:
            currency_obj = Currency.objects.filter(title__contains=item).first()
            if currency_obj:
                currencies[item] = currency_obj.pk
        payment_list = []
        count_dist = 0  # 分配记录数
        count_pay = 0  # 收款记录数
        errors = []  # 错误列表


        # 读入收款数据和处理数据
        for i in range(2, ws.max_row + 1):
            row = ws[i]
            # 第一步：读入并清理收款数据
            row_pay_dict = self.read_pay_record(row, field_list, bank_list)
            d = row_pay_dict
            # print('行号：%s, 内容：%s' % (i, row_pay_dict))
            if not row_pay_dict:
                break
            error = d.get('error')
            if error:
                if error.startswith('300'):
                    errors.append('行号：%s，错误：%s' % (i, error))
                # errors.append('行号：%s，错误：%s' % (i, error))
                continue
            print(row_pay_dict)


            # 第二步 创建收款记录
            # 检查订单记录是否已经存在, 如果不存在忽略本条记录，提示错误
            order_obj = ApplyOrder.objects.filter(order_number=d.get('order_number')).first()
            if not order_obj:
                errors.append('行号：%s，发票：%s 错误：%s' % (i, d.get('order_number'), '订单号未录入系统'))
                continue

            # 检查客户名是否为空
            if not order_obj.customer:
                errors.append('行号：%s，发票：%s 错误：%s' % (i, d.get('order_number'), '请补充订单的客户名'))
                continue

            # 检查reference号的缺失
            if not isinstance(d.get('reference'), int):
                errors.append('行号：%s，发票：%s 错误：%s' % (i, d.get('order_number'), '收款编号缺少'))
                continue

            #  检查收款记录是否存在
            pay_obj = Inwardpay.objects.filter(reference=d.get('reference')).first()
            # 不存在就新增一条收款记录 amount这一列填客户打款金额，另外加一列dist_amount
            got_amount = d.get('origin_pay') or d.get('got_amount')
            amount = d.get('amount') or got_amount
            if not pay_obj:
                pay_obj = Inwardpay(create_date=d.get('create_date'),
                                    bank_id=banks.get(d.get('bank')),
                                    currency_id=currencies.get(d.get('currency')),
                                    customer=order_obj.customer,
                                    amount=amount,
                                    got_amount=got_amount,
                                    torelate_amount=amount,
                                    remark=d.get('remark'),
                                    reference=d.get('reference'),
                                    confirm_status=3,
                                    status=1
                                    )

                payer_obj = Payer.objects.filter(title__icontains=d.get('payer')).first()
                if payer_obj:
                    pay_obj.payer = payer_obj
                print(pay_obj)
                pay_obj.save()
                count_pay += 1

            order_obj.rcvd_amount = sum(
                [item[0] for item in Pay2Orders.objects.filter(order=order_obj).values_list('amount')])
            order_obj.save()

            # 分配款项
            ## 查看分配记录是否存在，如果存在略过
            dist_obj = Pay2Orders.objects.filter(dist_ref=d.get('dist_ref')).first()
            if dist_obj:
                continue

            ## 检查是否超过订单的尾款金额
            dist_amount = d.get('dist_amount') or d.get('got_amount')
            dist_amount = Decimal(str(dist_amount))
            if order_obj.amount - order_obj.rcvd_amount - dist_amount < 0:
                errors.append('行号：%s，发票：%s 错误：%s' % (i, d.get('order_number'), '超过订单应收款'))
                continue

            ## 检查是否超过收款可分配金额
            if pay_obj.torelate_amount -dist_amount< 0:
                errors.append('行号：%s，发票：%s 错误：%s' % (i, d.get('order_number'), '超过收款可分配金额'))
                continue

            # 第三步 创建收款分配记录
            new_dist_obj = Pay2Orders(relate_date = d.get('create_date'),
                                      payment = pay_obj,
                                      order = order_obj,
                                      amount = dist_amount,
                                      dist_ref = d.get('dist_ref'),
                                      )

            new_dist_obj.save()
            # 扣减可关联金额
            pay_obj.torelate_amount = pay_obj.torelate_amount - dist_amount
            pay_obj.save()
            count_dist += 1

            # 刷新订单的总收款金额和应收款金额 values_list出来的是列表里面套tuple的结构
            order_obj.rcvd_amount =  order_obj.rcvd_amount + dist_amount
            order_obj.collect_amount = order_obj.amount - order_obj.rcvd_amount
            if order_obj.collect_amount < 0:
                errors.append('行号：%s，发票：%s 错误：%s' % (i, d.get('order_number'), '应收款不能为负数，请检查'))
                continue
            order_obj.save()

        # count分配记录数，count1收款记录数
        return count_dist, count_pay, errors

        # 上传之订单号预处理

    def parse_order_number(self, order_number):
        order_number = order_number.strip()
        type_choices = {item[1]: item[0] for item in ApplyOrder.type_choices}
        order_type = type_choices.get(order_number[0])
        sequence = order_number[1:5]
        order_split = order_number.split('-')
        sub_sequence = '0' if len(order_split) < 2 else order_split[-1]

        return order_type, sequence, sub_sequence

    def read_pay_record(self, row, field_list, bank_list):
        row_pay_dict = {}
        remark = '备注：'
        for n, field in field_list:
            # 避免读入空行
            if not row[0].value or not row[1].value:
                row_pay_dict['error'] = '400 空行'
                break
            # 默认的处理
            row_pay_dict[field] = row[n].value
            # 发票号, 原内容格式： 收汇 Jxxxx-2  印尼James  300错误需要提醒用户，400错误直接略过
            if field == "order_number":
                if not re.search('收汇', row[n].value):
                    row_pay_dict['error'] = '400 不是收汇'
                    break

                order_number = re.findall(Regex.order_number, row[n].value)
                if not order_number:
                    row_pay_dict['error'] = '400 发票号未找到'
                    break
                elif len(order_number) > 1:
                    row_pay_dict['error'] = '300 发现一个以上的发票号，请手动添加收款数据'
                    break
                else:
                    row_pay_dict[field] = order_number[0]
                    remark += row[n].value

            # 银行和货币
            elif field == "payer":
                row_pay_dict['bank'] = '广发'
                row_pay_dict['currency'] = '美元'
                for each in bank_list:
                    if each in row[n].value:
                        row_pay_dict['bank'] = each
                        break
                if '加拿大' in row[n].value:
                    row_pay_dict['currency'] = '加元'
                elif '人民币' in row[n].value:
                    row_pay_dict['currency'] = '人民币'

                # 提取付款公司名
                payer = re.findall(Regex.payer, row[n].value)
                if payer:
                    row_pay_dict[field] = ' '.join(payer[:2])
                remark += row[n].value
            # 提取分配金额
            elif field == "got_amount":
                try:
                    row_pay_dict[field] = Decimal(str(row[n].value))

                    remark = '%s %s' % (remark, row[n].value)
                except:
                    row_pay_dict['error'] = 'F列收款金额为空或者错误'
                    break

            # 需分配金额  原内容格式： "从 USD15866"
            elif field == "origin_pay":
                payment = 0
                # 这个地方加判断，因为origin_pay可能是空或者数值
                if isinstance(row[n].value, str) and re.search(Regex.pay_from_payment, row[n].value.strip()):
                    payment = re.search(Regex.pay_from_payment, row[n].value.strip()).group(1)
                    remark = '%s 从%s' % (remark, payment)
                row_pay_dict[field] = payment

        row_pay_dict['remark'] = remark

        return row_pay_dict

    # 下载上传用的模板文件
    def download(self, request, file_name, *args, **kwargs):
        print(request.POST, request.FILES, file_name)
        file_path = os.path.join(settings.MEDIA_ROOT, file_name)
        with open(file_path, 'rb') as f:
            try:
                response = HttpResponse(f)
                response['Content-Type'] = 'application/octet-stream'
                response['Content-Disposition'] = 'attachment;filename="%s"' % ('template_file.xlsx')

                return response
            except Exception as e:
                print(e)
                return HttpResponse("下载失败")

    #  下单
    def to_workshop(self, request, pk, *args, **kwargs):
        print(request.POST, )
        order_obj = self.model_class.objects.filter(pk=pk).first()
        if not order_obj:
            return HttpResponse('订单号不存在')

        if request.method == "GET":
            form = ConfirmApplyOrderModelForm(instance=order_obj, initial={'confirm_date': datetime.now()})
            return render(request, 'dipay/confirm_order.html', locals())
        else:
            form = ConfirmApplyOrderModelForm(request.POST, instance=order_obj)
            if form.is_valid():
                # 下单的动作1，把订单状态变为已下单
                form.instance.status = 2
                # 下单的动作2，把下单日期改为当日
                if not form.instance.confirm_date:
                    form.instance.confirm_date = datetime.now()
                # 下单的动作3，创建一条新的跟单记录, 此地要做一下判断，如果跟单记录已经存在就不要创建了。
                exists_follow_order = FollowOrder.objects.filter(order=order_obj).exists()
                if not exists_follow_order:
                    FollowOrder.objects.create(order=order_obj, salesman=order_obj.salesperson)

                form.save()
                order_number = order_obj.order_number
                link = reverse("stark:dipay_followorder_list")+"?q=%s" % order_number
                follow_order_link = "<a href='%s' target='_blank'>%s</a>" % (link, order_number)
                msg = '下单成功, 请将PI, 生产单，水单，唛头文件等邮件发到工厂跟单，已自动生成跟单记录:'+ follow_order_link
                msg = mark_safe(msg)
                return render(request, 'dipay/msg_after_submit.html', locals())
            else:
                return render(request, 'dipay/confirm_order.html', locals())

    # 手动创建订单
    def manual_add(self, request, *args, **kwargs):
        """ 手动创建订单  """
        if request.method == "GET":
            form = self.get_model_form("manual_add")()

            back_url = self.reverse_list_url()
            return render(request, "dipay/manual_add_new_order.html", locals())

        if request.method == "POST":
            form = self.get_model_form("manual_add")(data=request.POST)

            if form.is_valid():
                order_number = form.instance.order_number
                raw_data = [item.strip() for item in order_number.split('-')]
                if len(raw_data) == 1:
                    order_type = raw_data[0][0]
                    sequence = raw_data[0][1:5]
                    sub_sequence = 0
                    order_number = "%s%s" % (order_type, sequence)
                else:
                    order_number, sub_sequence = raw_data
                    order_type = order_number[0]
                    sequence = order_number[1:5]
                    order_number = "%s%s-%s" % (order_type, sequence, sub_sequence)
                # 获取订单类型列表，并转换为字典
                type_choices = {item[1]: item[0] for item in self.model_class.type_choices}

                form.instance.order_type = type_choices.get(order_type)
                form.instance.order_number = order_number
                form.instance.sequence = sequence
                form.instance.sub_sequence = sub_sequence
                try:
                    # 使用事务，避免数据不一致
                    with transaction.atomic():
                        apply_obj = form.save()
                        apply_obj.validate_unique()
                        # 同时创建跟单记录
                        follow_order_obj = FollowOrder(order=apply_obj, status=1, salesman=apply_obj.salesperson)
                        follow_order_obj.save()

                    msg = '手动创建%s成功，同时生成了跟单记录，请查看跟单表' % apply_obj
                except Exception as e:
                    print(e)
                    msg = '创建失败，可能已经存在同样的单号'
                return render(request, 'dipay/msg_after_submit.html', locals())
            else:
                return render(request, "dipay/manual_add_new_order.html", locals())


# 匹配模式的类
class Regex:
    # 提取订单号：
    order_number = r"([JMX]\d+-?\d)"
    # 提取金额
    amount = r'\d+[.]?\d*'
    # 提取付款公司名
    payer = r'[a-zA-Z]+'

    # 从收款表的最后一行提取金额，
    pay_from_payment = r"从\D*(\d+\.?\d*)"
